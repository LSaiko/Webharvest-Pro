"""
Excel Export Engine
Creates richly formatted Excel files with PowerBI/PivotTable analytics templates
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo


# Color palette
COLORS = {
    'primary_dark': '1A1A2E',
    'primary': '16213E',
    'accent': '0F3460',
    'highlight': 'E94560',
    'white': 'FFFFFF',
    'light_gray': 'F5F5F5',
    'mid_gray': 'CCCCCC',
    'dark_gray': '666666',
    'green': '27AE60',
    'blue': '2980B9',
    'orange': 'E67E22',
    'yellow': 'F1C40F',
}

def make_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def make_font(bold=False, color='000000', size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Arial')

def make_border(style='thin'):
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)

def make_center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def make_left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, cols, bg_color=None, font_color='FFFFFF', font_size=11):
    bg = bg_color or COLORS['primary_dark']
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(bg)
        cell.font = make_font(bold=True, color=font_color, size=font_size)
        cell.alignment = make_center()
        cell.border = make_border()


def auto_fit_columns(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def create_excel_export(all_results: list, output_path: str) -> str:
    """
    Create a comprehensive Excel workbook from scrape results.
    Includes: Summary, Raw Data, Tables, Links, Analytics Template, PowerBI Guide
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _create_summary_sheet(wb, all_results)
    _create_raw_data_sheet(wb, all_results)
    _create_tables_sheet(wb, all_results)
    _create_links_sheet(wb, all_results)
    _create_analytics_template_sheet(wb, all_results)
    _create_powerbi_guide_sheet(wb)
    _create_pivot_guide_sheet(wb)
    _create_changelog_sheet(wb, all_results)

    wb.save(output_path)
    return output_path


def _create_summary_sheet(wb, all_results):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_properties.tabColor = COLORS['highlight']

    # Title block
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "🌐 Web Scraper — Summary Dashboard"
    title_cell.font = make_font(bold=True, color='FFFFFF', size=16)
    title_cell.fill = make_fill(COLORS['primary_dark'])
    title_cell.alignment = make_center()
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Sites: {len(all_results)}"
    sub.font = make_font(italic=True, color='FFFFFF', size=10)
    sub.fill = make_fill(COLORS['accent'])
    sub.alignment = make_center()
    ws.row_dimensions[2].height = 22

    # Stats cards row
    stats_row = 4
    successful = [r for r in all_results if r.get('success')]
    failed = [r for r in all_results if not r.get('success')]
    total_tables = sum(len(r.get('data', {}).get('tables', [])) for r in successful)
    total_links = sum(len(r.get('data', {}).get('links', [])) for r in successful)

    stats = [
        ('Total URLs', len(all_results), COLORS['accent']),
        ('Successful', len(successful), COLORS['green']),
        ('Failed / Blocked', len(failed), COLORS['highlight']),
        ('Tables Found', total_tables, COLORS['blue']),
        ('Links Collected', total_links, COLORS['orange']),
    ]

    for i, (label, val, color) in enumerate(stats):
        col = i * 2 + 1
        ws.merge_cells(start_row=stats_row, start_column=col, end_row=stats_row, end_column=col+1)
        ws.merge_cells(start_row=stats_row+1, start_column=col, end_row=stats_row+1, end_column=col+1)
        label_cell = ws.cell(row=stats_row, column=col, value=label)
        label_cell.fill = make_fill(color)
        label_cell.font = make_font(bold=True, color='FFFFFF', size=10)
        label_cell.alignment = make_center()
        ws.row_dimensions[stats_row].height = 22
        val_cell = ws.cell(row=stats_row+1, column=col, value=val)
        val_cell.fill = make_fill(COLORS['light_gray'])
        val_cell.font = make_font(bold=True, size=18)
        val_cell.alignment = make_center()
        ws.row_dimensions[stats_row+1].height = 36

    # Per-site table
    table_start = 8
    headers = ['#', 'URL', 'Status', 'Title', 'Paragraphs', 'Tables', 'Links', 'Scraped At', 'Content Hash']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=c, value=h)
    style_header_row(ws, table_start, len(headers), bg_color=COLORS['primary'])

    for i, r in enumerate(all_results):
        row = table_start + 1 + i
        data = r.get('data') or {}
        status = '✅ OK' if r.get('success') else f"❌ {r.get('error', 'Error')[:40]}"
        row_data = [
            i + 1,
            r['url'],
            status,
            data.get('title', '')[:80],
            len(data.get('paragraphs', [])),
            len(data.get('tables', [])),
            len(data.get('links', [])),
            r.get('timestamp', '')[:19],
            r.get('meta', {}).get('content_hash', '')
        ]
        fill = make_fill('F0FFF4') if r.get('success') else make_fill('FFF0F0')
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = fill
            cell.alignment = make_left()
            cell.border = make_border('thin')
            cell.font = make_font(size=10)
        ws.row_dimensions[row].height = 18

    # Add table style
    if all_results:
        last_row = table_start + len(all_results)
        last_col = get_column_letter(len(headers))
        tbl = Table(displayName="SummaryTable", ref=f"A{table_start}:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    auto_fit_columns(ws)
    ws.freeze_panes = f'A{table_start + 1}'


def _create_raw_data_sheet(wb, all_results):
    ws = wb.create_sheet("📄 Raw Data")
    ws.sheet_properties.tabColor = COLORS['blue']

    headers = ['Source URL', 'Section Type', 'Level/Index', 'Content', 'Scraped At']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers), bg_color=COLORS['primary'])

    row = 2
    for r in all_results:
        if not r.get('success'):
            continue
        data = r.get('data', {})
        url = r['url']
        ts = r.get('timestamp', '')[:19]

        # Headings
        for h in data.get('headings', []):
            ws.append([url, 'Heading', h['level'].upper(), h['text'], ts])
            ws.row_dimensions[row].height = 16
            row += 1

        # Paragraphs
        for i, p in enumerate(data.get('paragraphs', [])[:50]):
            ws.append([url, 'Paragraph', str(i+1), p, ts])
            ws.row_dimensions[row].height = 16
            row += 1

        # List items
        for li, lst in enumerate(data.get('lists', [])):
            for item in lst:
                ws.append([url, 'List Item', str(li+1), item, ts])
                row += 1

        # Description
        if data.get('description'):
            ws.append([url, 'Meta Description', '', data['description'], ts])
            row += 1

    # Format
    for r_idx in range(2, row):
        for c in range(1, 6):
            cell = ws.cell(row=r_idx, column=c)
            cell.font = make_font(size=10)
            cell.alignment = make_left()
            cell.border = make_border('hair')
            if r_idx % 2 == 0:
                cell.fill = make_fill('F8F9FA')

    if row > 2:
        last_col = get_column_letter(len(headers))
        tbl = Table(displayName="RawDataTable", ref=f"A1:{last_col}{row-1}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tbl)

    auto_fit_columns(ws, min_width=15, max_width=80)
    ws.freeze_panes = 'A2'


def _create_tables_sheet(wb, all_results):
    ws = wb.create_sheet("📋 Tables")
    ws.sheet_properties.tabColor = COLORS['green']

    row = 1
    table_num = 0

    for r in all_results:
        if not r.get('success'):
            continue
        for t in r.get('data', {}).get('tables', []):
            table_num += 1

            # Table header banner
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            banner = ws.cell(row=row, column=1,
                value=f"TABLE {table_num} — Source: {r['url'][:80]}")
            banner.fill = make_fill(COLORS['accent'])
            banner.font = make_font(bold=True, color='FFFFFF', size=10)
            banner.alignment = make_left()
            ws.row_dimensions[row].height = 20
            row += 1

            # Column headers
            headers = t.get('headers', [])
            table_rows = t.get('rows', [])
            if not headers and table_rows:
                headers = table_rows[0]
                table_rows = table_rows[1:]

            max_cols = max((len(headers), max((len(r2) for r2 in table_rows), default=0)), default=1)

            if headers:
                for c, h in enumerate(headers[:20], 1):
                    cell = ws.cell(row=row, column=c, value=h)
                    cell.fill = make_fill(COLORS['primary_dark'])
                    cell.font = make_font(bold=True, color='FFFFFF', size=10)
                    cell.alignment = make_center()
                    cell.border = make_border()
                ws.row_dimensions[row].height = 18
                row += 1

            for data_row in table_rows[:100]:
                for c, val in enumerate(data_row[:20], 1):
                    cell = ws.cell(row=row, column=c, value=val)
                    cell.font = make_font(size=10)
                    cell.alignment = make_left()
                    cell.border = make_border('hair')
                    if row % 2 == 0:
                        cell.fill = make_fill('F0F8FF')
                ws.row_dimensions[row].height = 16
                row += 1

            row += 1  # Gap between tables

    if row == 1:
        ws['A1'] = 'No tables found in scraped pages.'
        ws['A1'].font = make_font(italic=True, color=COLORS['dark_gray'])

    auto_fit_columns(ws)


def _create_links_sheet(wb, all_results):
    ws = wb.create_sheet("🔗 Links")
    ws.sheet_properties.tabColor = COLORS['orange']

    headers = ['Source URL', 'Link Text', 'Link URL', 'External?', 'Domain', 'Scraped At']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers), bg_color=COLORS['primary'])

    row = 2
    for r in all_results:
        if not r.get('success'):
            continue
        ts = r.get('timestamp', '')[:19]
        for link in r.get('data', {}).get('links', []):
            domain = urlparse_domain(link.get('url', ''))
            is_ext = link.get('external', False)
            ws.append([
                r['url'],
                link.get('text', '')[:100],
                link.get('url', ''),
                'Yes' if is_ext else 'No',
                domain,
                ts
            ])
            cell_fill = make_fill('FFF5EE') if is_ext else make_fill('F0FFF0')
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.font = make_font(size=10)
                cell.alignment = make_left()
                cell.fill = cell_fill
                cell.border = make_border('hair')
            row += 1

    if row > 2:
        last_col = get_column_letter(len(headers))
        tbl = Table(displayName="LinksTable", ref=f"A1:{last_col}{row-1}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)

    auto_fit_columns(ws, max_width=60)
    ws.freeze_panes = 'A2'


def urlparse_domain(url):
    try:
        return urlparse(url).netloc
    except Exception:
        return ''


def _create_analytics_template_sheet(wb, all_results):
    ws = wb.create_sheet("📈 Analytics Template")
    ws.sheet_properties.tabColor = COLORS['highlight']

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = '📈 Analytics Template — Pre-built Formulas for Dashboard Creation'
    ws['A1'].font = make_font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = make_fill(COLORS['highlight'])
    ws['A1'].alignment = make_center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:J2')
    ws['A2'] = 'Use these formulas and data structures as starting points for PivotTables, PowerBI, or Tableau dashboards'
    ws['A2'].font = make_font(italic=True, size=10)
    ws['A2'].fill = make_fill('FFF3CD')
    ws['A2'].alignment = make_center()

    # KPI metrics section
    row = 4
    ws.merge_cells(f'A{row}:J{row}')
    ws.cell(row=row, column=1, value='📊 KEY METRICS — AUTO-CALCULATED FROM RAW DATA SHEET')
    ws.cell(row=row, column=1).font = make_font(bold=True, color='FFFFFF', size=11)
    ws.cell(row=row, column=1).fill = make_fill(COLORS['accent'])
    ws.row_dimensions[row].height = 22

    row += 1
    kpis = [
        ('Total Records', "=COUNTA('📄 Raw Data'!A:A)-1", 'Count of all scraped content rows'),
        ('Unique Sources', "=SUMPRODUCT(1/COUNTIF('📄 Raw Data'!A2:A1000,'📄 Raw Data'!A2:A1000))", 'Distinct source URLs'),
        ('Heading Count', "=COUNTIF('📄 Raw Data'!B:B,\"Heading\")", 'Total headings scraped'),
        ('Paragraph Count', "=COUNTIF('📄 Raw Data'!B:B,\"Paragraph\")", 'Total paragraphs scraped'),
        ('External Links', "=COUNTIF('🔗 Links'!D:D,\"Yes\")", 'Count of outbound external links'),
        ('Internal Links', "=COUNTIF('🔗 Links'!D:D,\"No\")", 'Count of internal page links'),
        ('Tables Scraped', "=COUNTA('📋 Tables'!A:A)", 'Rows of table data collected'),
        ('Success Rate', "=COUNTIF('📊 Summary'!C:C,\"✅ OK\")/COUNTA('📊 Summary'!A:A)", 'Scrape success rate'),
    ]

    metric_headers = ['Metric Name', 'Formula / Value', 'Description', 'Result']
    for c, h in enumerate(metric_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(metric_headers), bg_color=COLORS['primary_dark'])
    row += 1

    for name, formula, desc in kpis:
        ws.cell(row=row, column=1, value=name).font = make_font(bold=True, size=10)
        formula_cell = ws.cell(row=row, column=2, value=formula)
        formula_cell.font = Font(name='Courier New', size=9, color='0000FF')
        formula_cell.fill = make_fill('E8F4FD')
        ws.cell(row=row, column=3, value=desc).font = make_font(italic=True, size=9, color=COLORS['dark_gray'])
        ws.cell(row=row, column=4, value='← Activate formula in col B').font = make_font(italic=True, size=9, color='888888')
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = make_border('hair')
            ws.cell(row=row, column=c).alignment = make_left()
        if row % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = make_fill('F8F9FA')
        row += 1

    # Pivot Table instructions
    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws.cell(row=row, column=1, value='🔄 PIVOT TABLE RECIPES — Copy these instructions to create instant dashboards')
    ws.cell(row=row, column=1).font = make_font(bold=True, color='FFFFFF', size=11)
    ws.cell(row=row, column=1).fill = make_fill(COLORS['green'])
    ws.row_dimensions[row].height = 22
    row += 1

    pivot_recipes = [
        ('Content by Source', 'Raw Data sheet', 'Rows: Source URL | Values: COUNT of Content | Filter: Section Type', 'Shows how much content was collected per website'),
        ('Content Type Breakdown', 'Raw Data sheet', 'Rows: Section Type | Values: COUNT of Content', 'Pie/bar chart of content categories (Heading/Para/List)'),
        ('Link Analysis', 'Links sheet', 'Rows: Domain | Values: COUNT of Link URL | Filter: External?=Yes', 'Most referenced external domains'),
        ('Scrape Timeline', 'Summary sheet', 'Rows: Scraped At (Group by Day/Hour) | Values: COUNT of URLs', 'Monitor scrape frequency over time'),
        ('Table Data Density', 'Summary sheet', 'Rows: URL | Values: SUM of Tables + SUM of Links', 'Which sites have richest structured data'),
        ('Success vs Failure', 'Summary sheet', 'Rows: Status | Values: COUNT of #', 'Donut chart of scrape success rates'),
    ]

    pivot_headers = ['Pivot Name', 'Data Source Sheet', 'Configuration', 'Purpose / Chart Type']
    for c, h in enumerate(pivot_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(pivot_headers), bg_color=COLORS['accent'])
    row += 1

    for i, (name, source, config, purpose) in enumerate(pivot_recipes):
        ws.cell(row=row, column=1, value=name).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=2, value=source).font = make_font(size=10)
        ws.cell(row=row, column=3, value=config).font = Font(name='Courier New', size=9, color='0000FF')
        ws.cell(row=row, column=3).fill = make_fill('EFF7FF')
        ws.cell(row=row, column=4, value=purpose).font = make_font(italic=True, size=9)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = make_border('hair')
            ws.cell(row=row, column=c).alignment = make_left()
        row += 1

    auto_fit_columns(ws, min_width=15, max_width=70)


def _create_powerbi_guide_sheet(wb):
    ws = wb.create_sheet("⚡ PowerBI Guide")
    ws.sheet_properties.tabColor = '2980B9'

    ws.merge_cells('A1:H1')
    ws['A1'] = '⚡ PowerBI Integration Guide'
    ws['A1'].font = make_font(bold=True, color='FFFFFF', size=16)
    ws['A1'].fill = make_fill('2980B9')
    ws['A1'].alignment = make_center()
    ws.row_dimensions[1].height = 40

    sections = [
        ('🔌 CONNECTING THIS FILE TO POWERBI', 'step_header', [
            ('Step 1', 'Open PowerBI Desktop → Home → Get Data → Excel Workbook'),
            ('Step 2', "Browse to this file and click 'Open'"),
            ('Step 3', "Select sheets: '📊 Summary', '📄 Raw Data', '🔗 Links' → Load"),
            ('Step 4', "In the Data pane, select '📄 Raw Data' table"),
            ('Step 5', "Build relationships: connect 'Source URL' across all tables"),
        ]),
        ('📊 RECOMMENDED VISUALS TO BUILD', 'step_header', [
            ('Card Visual', "Drag 'Total Records' from Analytics Template → KPI metric tile"),
            ('Bar Chart', "X-axis: Source URL | Y-axis: COUNT of Content | Filter: Section Type = Heading"),
            ('Donut Chart', "Values: COUNT of Status from Summary | Legend: Status (Success/Failed)"),
            ('Table/Matrix', "Rows: Domain | Values: COUNT of Links — shows top linked sites"),
            ('Line Chart', "X-axis: Scraped At | Y-axis: COUNT of rows — monitor scrape history over time"),
            ('Word Cloud', "Feed 'Content' column from Raw Data into a Word Cloud custom visual"),
        ]),
        ('🔧 DAX FORMULAS FOR POWERBI', 'formula_header', [
            ('Success Rate %', "Success Rate = DIVIDE(COUNTROWS(FILTER('📊 Summary', '📊 Summary'[Status] = \"✅ OK\")), COUNTROWS('📊 Summary'))"),
            ('Avg Content/Site', "Avg Content = DIVIDE(COUNTROWS('📄 Raw Data'), DISTINCTCOUNT('📄 Raw Data'[Source URL]))"),
            ('External Link %', "Ext Link % = DIVIDE(COUNTROWS(FILTER('🔗 Links', '🔗 Links'[External?]=\"Yes\")), COUNTROWS('🔗 Links'))"),
            ('Last Scraped', "Last Update = MAXX('📊 Summary', '📊 Summary'[Scraped At])"),
            ('Heading Density', "Heading Density = DIVIDE(COUNTROWS(FILTER('📄 Raw Data', '📄 Raw Data'[Section Type]=\"Heading\")), DISTINCTCOUNT('📄 Raw Data'[Source URL]))"),
        ]),
        ('🎨 DASHBOARD LAYOUT SUGGESTIONS', 'step_header', [
            ('Page 1: Overview', "4 KPI cards (Sites, Records, Success Rate, Last Updated) + Donut + Bar chart"),
            ('Page 2: Content', "Word cloud + Table of top paragraphs + Content type breakdown by source"),
            ('Page 3: Links', "Link domain bar chart + External vs Internal ratio + Top 10 linked domains"),
            ('Page 4: Monitor', "Timeline line chart showing scraping history + Status over time"),
        ]),
    ]

    row = 3
    for section_title, section_type, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        hdr = ws.cell(row=row, column=1, value=section_title)
        hdr.font = make_font(bold=True, color='FFFFFF', size=12)
        hdr.fill = make_fill('1A5276' if section_type == 'step_header' else '145A32')
        hdr.alignment = make_left()
        ws.row_dimensions[row].height = 25
        row += 1

        for label, content in items:
            ws.cell(row=row, column=1, value=label).font = make_font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = make_fill('D6EAF8')
            ws.cell(row=row, column=1).alignment = make_left()

            content_cell = ws.cell(row=row, column=2, value=content)
            if section_type == 'formula_header':
                content_cell.font = Font(name='Courier New', size=9, color='0000FF')
                content_cell.fill = make_fill('E8F8F5')
            else:
                content_cell.font = make_font(size=10)
                content_cell.fill = make_fill('EBF5FB')
            content_cell.alignment = make_left()
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

            for c in range(1, 9):
                ws.cell(row=row, column=c).border = make_border('hair')
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 90


def _create_pivot_guide_sheet(wb):
    ws = wb.create_sheet("🔄 Excel PivotTable Guide")

    ws.merge_cells('A1:F1')
    ws['A1'] = '🔄 Excel PivotTable & Advanced Analytics Guide'
    ws['A1'].font = make_font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = make_fill(COLORS['green'])
    ws['A1'].alignment = make_center()
    ws.row_dimensions[1].height = 35

    guides = [
        ('CREATE A PIVOT TABLE', [
            ('1', 'Click any cell in the "📄 Raw Data" sheet'),
            ('2', 'Insert tab → PivotTable → New Worksheet'),
            ('3', 'Drag "Source URL" to Rows field'),
            ('4', 'Drag "Section Type" to Columns field'),
            ('5', 'Drag "Content" to Values field → set to COUNT'),
            ('6', 'Result: Grid showing content type breakdown per site'),
        ]),
        ('USEFUL EXCEL FORMULAS', [
            ('COUNTIF', '=COUNTIF(B:B,"Heading") — Count specific content types'),
            ('SUMPRODUCT', '=SUMPRODUCT((A2:A1000=A2)*1) — Count entries per URL'),
            ('UNIQUE', '=UNIQUE(A2:A1000) — List all unique source URLs (Excel 365)'),
            ('FILTER', '=FILTER(C:C,B:B="Heading") — Show only headings'),
            ('SORT', '=SORT(UNIQUE(E:E),1,1) — Alphabetical domain list'),
            ('TEXTJOIN', '=TEXTJOIN(", ",TRUE,IF(A:A=A2,C:C,"")) — Combine content per URL'),
        ]),
        ('CHART RECOMMENDATIONS', [
            ('Content Volume', 'Bar chart: URLs on X-axis, row counts on Y-axis'),
            ('Link Types', 'Stacked bar: Internal vs External links per domain'),
            ('Success Rate', 'Donut chart: Success vs Failed scrapes'),
            ('Content Mix', 'Pie chart: Headings vs Paragraphs vs Lists'),
            ('Timeline', 'Line chart: Scrapes over time from Changelog sheet'),
        ]),
        ('SLICERS & FILTERS', [
            ('Add Slicer', 'Click PivotTable → PivotTable Analyze → Insert Slicer'),
            ('Section Type', 'Add slicer for Section Type to toggle Heading/Para/List'),
            ('Source URL', 'Add slicer for Source URL to isolate specific sites'),
            ('Date Filter', 'Use Timeline slicer on Scraped At column for time filtering'),
        ]),
    ]

    row = 3
    for section, items in guides:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        hdr = ws.cell(row=row, column=1, value=f'  {section}')
        hdr.font = make_font(bold=True, color='FFFFFF', size=11)
        hdr.fill = make_fill(COLORS['green'])
        ws.row_dimensions[row].height = 22
        row += 1

        for step, desc in items:
            ws.cell(row=row, column=1, value=step).font = make_font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = make_fill('D5F5E3')
            ws.cell(row=row, column=1).alignment = make_center()
            desc_cell = ws.cell(row=row, column=2, value=desc)
            if '=' in desc:
                desc_cell.font = Font(name='Courier New', size=9, color='0000FF')
                desc_cell.fill = make_fill('E8F8F5')
            else:
                desc_cell.font = make_font(size=10)
                desc_cell.fill = make_fill('EAFAF1')
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            desc_cell.alignment = make_left()
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = make_border('hair')
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 80


def _create_changelog_sheet(wb, all_results):
    ws = wb.create_sheet("📝 Changelog")
    ws.sheet_properties.tabColor = COLORS['dark_gray']

    ws.merge_cells('A1:E1')
    ws['A1'] = '📝 Scrape History & Changelog'
    ws['A1'].font = make_font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = make_fill(COLORS['primary_dark'])
    ws['A1'].alignment = make_center()
    ws.row_dimensions[1].height = 30

    headers = ['Run Timestamp', 'URL', 'Status', 'Records Collected', 'Content Hash']
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))

    for i, r in enumerate(all_results):
        row = i + 3
        data = r.get('data') or {}
        total_records = (len(data.get('headings', [])) + len(data.get('paragraphs', [])) +
                        len(data.get('links', [])))
        ws.append([
            r.get('timestamp', '')[:19],
            r['url'],
            '✅ Success' if r.get('success') else f"❌ {r.get('error', '')[:40]}",
            total_records,
            r.get('meta', {}).get('content_hash', 'N/A')
        ])
        fill = make_fill('F0FFF4') if r.get('success') else make_fill('FFF0F0')
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.font = make_font(size=10)
            cell.fill = fill
            cell.border = make_border('hair')
            cell.alignment = make_left()

    auto_fit_columns(ws)
    ws.freeze_panes = 'A3'
